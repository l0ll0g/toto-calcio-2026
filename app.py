from flask import Flask, render_template, request, jsonify, session, send_file
import os, hashlib, json, io, secrets, string, re
from functools import wraps
from datetime import datetime, timezone, timedelta
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.middleware.proxy_fix import ProxyFix
import logging

app = Flask(__name__)
# Dietro un reverse proxy/hosting (Render, Railway, nginx): fidati degli header
# X-Forwarded-* così Flask sa che la connessione esterna è HTTPS e l'IP reale.
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1)

# SECRET_KEY: in produzione DEVE arrivare da variabile d'ambiente.
# In sviluppo, se assente, ne generiamo una casuale (le sessioni non
# sopravvivono al riavvio, ma è sicura).
_secret = os.environ.get('SECRET_KEY')
if not _secret:
    _secret = secrets.token_hex(32)
    print('[WARN] SECRET_KEY non impostata: ne uso una casuale temporanea. '
          'In produzione imposta la variabile SECRET_KEY.')
app.secret_key = _secret
logging.basicConfig(level=logging.INFO,
                    format='%(asctime)s [%(levelname)s] %(message)s')

# Sicurezza cookie di sessione
_is_prod = os.environ.get('FLASK_ENV') == 'production' or os.environ.get('PRODUCTION') == '1'
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE='Lax',
    SESSION_COOKIE_SECURE=_is_prod,     # True solo dietro HTTPS in produzione
    PERMANENT_SESSION_LIFETIME=timedelta(days=30),
    MAX_CONTENT_LENGTH=1 * 1024 * 1024, # max 1 MB per richiesta
)
EMAIL_RE = re.compile(r'^[^@\s]+@[^@\s]+\.[^@\s]+$')

# ── Config ─────────────────────────────────────────────────────────────
ADMIN_EMAIL  = 'lorenzogucci05@gmail.com'
WC_DEADLINE  = datetime(2026, 6, 11, 20, 59, 59, tzinfo=timezone.utc)
WC_GROUPS    = ['A','B','C','D','E','F','G','H','I','J','K','L']
API_FOOTBALL_KEY = os.environ.get('API_FOOTBALL_KEY', '')

# ── Live results: Highlightly (soccer.highlightly.net) ─────────────────────
# Chiave gratuita su https://highlightly.net/login (100 richieste/giorno).
# Il piano gratuito copre amichevoli internazionali (lega "Friendlies") e i
# campionati principali. Impostala come variabile d'ambiente HIGHLIGHTLY_KEY.
HIGHLIGHTLY_KEY  = os.environ.get('HIGHLIGHTLY_KEY', '')
HIGHLIGHTLY_BASE = 'https://soccer.highlightly.net'
HIGHLIGHTLY_HOST = os.environ.get('HIGHLIGHTLY_HOST', 'soccer.highlightly.net').strip() or 'soccer.highlightly.net'
# Cloudflare (davanti a Highlightly) blocca le richieste "da bot" con error 1010.
# Aggiungendo uno User-Agent da browser le richieste passano (verificato da Render).
def _hl_headers():
    return {
        'x-rapidapi-key': HIGHLIGHTLY_KEY,
        'x-rapidapi-host': HIGHLIGHTLY_HOST,
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36',
        'Accept': 'application/json',
    }
# Quale lega interrogare:
#  - "9294"  = Amichevoli internazionali (per i test prima del Mondiale)
#  - id lega FIFA World Cup = per il Mondiale (da impostare quando disponibile)
# Configurabile via env LIVE_LEAGUE_ID. Default amichevoli per ora.
LIVE_LEAGUE_ID   = os.environ.get('LIVE_LEAGUE_ID', '9294').strip() or '9294'
LIVE_SEASON      = os.environ.get('LIVE_SEASON', '2026').strip() or '2026'
LIVE_POLL_SECONDS = int(os.environ.get('LIVE_POLL_SECONDS', '120'))  # cache lato server (default 2 min, protegge il limite 100 richieste/giorno)
# (compat: vecchie variabili, non più usate dal motore ma non rompono nulla)
FOOTBALL_DATA_KEY = os.environ.get('FOOTBALL_DATA_KEY', '')
# Map: nome squadra nella nostra app -> nome su Highlightly (inglese)
LIVE_TEAM_ALIASES = {
    'Corea del Sud':'South Korea','Repubblica Ceca':'Czech Republic','Sudafrica':'South Africa',
    'Messico':'Mexico','Germania':'Germany','Spagna':'Spain','Francia':'France','Inghilterra':'England',
    'Brasile':'Brazil','Croazia':'Croatia','Belgio':'Belgium','Svizzera':'Switzerland','Olanda':'Netherlands',
    'Giappone':'Japan','Marocco':'Morocco','Senegal':'Senegal','Norvegia':'Norway','Arabia Saudita':'Saudi Arabia',
    'Capo Verde':'Cape Verde','Uzbekistan':'Uzbekistan','Giordania':'Jordan','Egitto':'Egypt','Nuova Zelanda':'New Zealand',
    'Algeria':'Algeria','Austria':'Austria','Portogallo':'Portugal','Colombia':'Colombia','RD Congo':'DR Congo',
    'Ghana':'Ghana','Panama':'Panama','Uruguay':'Uruguay','Iran':'Iran','Iraq':'Iraq','Tunisia':'Tunisia',
    'Svezia':'Sweden','Ecuador':'Ecuador',"Costa d'Avorio":'Ivory Coast','Curacao':'Curaçao','Haiti':'Haiti',
    'Scozia':'Scotland','Canada':'Canada','USA':'USA','Paraguay':'Paraguay','Australia':'Australia',
    'Turchia':'Turkey','Qatar':'Qatar','Bosnia':'Bosnia and Herzegovina','Argentina':'Argentina',
    'Ungheria':'Hungary','Finlandia':'Finland','Cile':'Chile','Slovenia':'Slovenia','Ucraina':'Ukraine',
    'Italia':'Italy','Grecia':'Greece','Irlanda del Nord':'Northern Ireland','Kazakistan':'Kazakhstan',
    'Costa Rica':'Costa Rica','Nigeria':'Nigeria',
    'Islanda':'Iceland','Venezuela':'Venezuela','Pakistan':'Pakistan','Afghanistan':'Afghanistan',
}
LIVE_ENABLED = bool(HIGHLIGHTLY_KEY)

# Match labels for Excel (id -> home/away)
WC_MATCH_SCHEDULE = {
  'A':[{'id':'wc-A-m1','home':'Messico','away':'Sudafrica'},{'id':'wc-A-m2','home':'Corea del Sud','away':'Repubblica Ceca'},{'id':'wc-A-m3','home':'Repubblica Ceca','away':'Sudafrica'},{'id':'wc-A-m4','home':'Messico','away':'Corea del Sud'},{'id':'wc-A-m5','home':'Sudafrica','away':'Corea del Sud'},{'id':'wc-A-m6','home':'Repubblica Ceca','away':'Messico'}],
  'B':[{'id':'wc-B-m1','home':'Canada','away':'Bosnia'},{'id':'wc-B-m2','home':'Svizzera','away':'Qatar'},{'id':'wc-B-m3','home':'Svizzera','away':'Bosnia'},{'id':'wc-B-m4','home':'Canada','away':'Qatar'},{'id':'wc-B-m5','home':'Svizzera','away':'Canada'},{'id':'wc-B-m6','home':'Bosnia','away':'Qatar'}],
  'C':[{'id':'wc-C-m1','home':'Brasile','away':'Marocco'},{'id':'wc-C-m2','home':'Haiti','away':'Scozia'},{'id':'wc-C-m3','home':'Scozia','away':'Marocco'},{'id':'wc-C-m4','home':'Brasile','away':'Haiti'},{'id':'wc-C-m5','home':'Marocco','away':'Haiti'},{'id':'wc-C-m6','home':'Scozia','away':'Brasile'}],
  'D':[{'id':'wc-D-m1','home':'USA','away':'Paraguay'},{'id':'wc-D-m2','home':'Australia','away':'Turchia'},{'id':'wc-D-m3','home':'Turchia','away':'Paraguay'},{'id':'wc-D-m4','home':'USA','away':'Australia'},{'id':'wc-D-m5','home':'Turchia','away':'USA'},{'id':'wc-D-m6','home':'Paraguay','away':'Australia'}],
  'E':[{'id':'wc-E-m1','home':'Germania','away':'Curacao'},{'id':'wc-E-m2','home':"Costa d'Avorio",'away':'Ecuador'},{'id':'wc-E-m3','home':'Germania','away':"Costa d'Avorio"},{'id':'wc-E-m4','home':'Ecuador','away':'Curacao'},{'id':'wc-E-m5','home':'Curacao','away':"Costa d'Avorio"},{'id':'wc-E-m6','home':'Ecuador','away':'Germania'}],
  'F':[{'id':'wc-F-m1','home':'Olanda','away':'Giappone'},{'id':'wc-F-m2','home':'Svezia','away':'Tunisia'},{'id':'wc-F-m3','home':'Tunisia','away':'Giappone'},{'id':'wc-F-m4','home':'Olanda','away':'Svezia'},{'id':'wc-F-m5','home':'Tunisia','away':'Olanda'},{'id':'wc-F-m6','home':'Giappone','away':'Svezia'}],
  'G':[{'id':'wc-G-m1','home':'Belgio','away':'Egitto'},{'id':'wc-G-m2','home':'Iran','away':'Nuova Zelanda'},{'id':'wc-G-m3','home':'Belgio','away':'Iran'},{'id':'wc-G-m4','home':'Nuova Zelanda','away':'Egitto'},{'id':'wc-G-m5','home':'Nuova Zelanda','away':'Belgio'},{'id':'wc-G-m6','home':'Egitto','away':'Iran'}],
  'H':[{'id':'wc-H-m1','home':'Spagna','away':'Capo Verde'},{'id':'wc-H-m2','home':'Arabia Saudita','away':'Uruguay'},{'id':'wc-H-m3','home':'Spagna','away':'Arabia Saudita'},{'id':'wc-H-m4','home':'Uruguay','away':'Capo Verde'},{'id':'wc-H-m5','home':'Capo Verde','away':'Arabia Saudita'},{'id':'wc-H-m6','home':'Uruguay','away':'Spagna'}],
  'I':[{'id':'wc-I-m1','home':'Francia','away':'Senegal'},{'id':'wc-I-m2','home':'Iraq','away':'Norvegia'},{'id':'wc-I-m3','home':'Francia','away':'Iraq'},{'id':'wc-I-m4','home':'Norvegia','away':'Senegal'},{'id':'wc-I-m5','home':'Norvegia','away':'Francia'},{'id':'wc-I-m6','home':'Senegal','away':'Iraq'}],
  'J':[{'id':'wc-J-m1','home':'Austria','away':'Giordania'},{'id':'wc-J-m2','home':'Argentina','away':'Algeria'},{'id':'wc-J-m3','home':'Argentina','away':'Austria'},{'id':'wc-J-m4','home':'Giordania','away':'Algeria'},{'id':'wc-J-m5','home':'Algeria','away':'Austria'},{'id':'wc-J-m6','home':'Giordania','away':'Argentina'}],
  'K':[{'id':'wc-K-m1','home':'Portogallo','away':'RD Congo'},{'id':'wc-K-m2','home':'Uzbekistan','away':'Colombia'},{'id':'wc-K-m3','home':'Portogallo','away':'Uzbekistan'},{'id':'wc-K-m4','home':'Colombia','away':'RD Congo'},{'id':'wc-K-m5','home':'Colombia','away':'Portogallo'},{'id':'wc-K-m6','home':'RD Congo','away':'Uzbekistan'}],
  'L':[{'id':'wc-L-m1','home':'Inghilterra','away':'Croazia'},{'id':'wc-L-m2','home':'Ghana','away':'Panama'},{'id':'wc-L-m3','home':'Inghilterra','away':'Ghana'},{'id':'wc-L-m4','home':'Panama','away':'Croazia'},{'id':'wc-L-m5','home':'Panama','away':'Inghilterra'},{'id':'wc-L-m6','home':'Croazia','away':'Ghana'}],
}

# Date (UTC) delle partite del Mondiale: usate per cercare i risultati PER DATA,
# esattamente come per le amichevoli (la fonte abbina per coppia di squadre).
WC_MATCH_DATES = {'wc-A-m1':'2026-06-11', 'wc-A-m2':'2026-06-12', 'wc-A-m3':'2026-06-18', 'wc-A-m4':'2026-06-19', 'wc-A-m5':'2026-06-25', 'wc-A-m6':'2026-06-25', 'wc-B-m1':'2026-06-12', 'wc-B-m2':'2026-06-13', 'wc-B-m3':'2026-06-18', 'wc-B-m4':'2026-06-19', 'wc-B-m5':'2026-06-24', 'wc-B-m6':'2026-06-24', 'wc-C-m1':'2026-06-14', 'wc-C-m2':'2026-06-14', 'wc-C-m3':'2026-06-20', 'wc-C-m4':'2026-06-20', 'wc-C-m5':'2026-06-25', 'wc-C-m6':'2026-06-25', 'wc-D-m1':'2026-06-13', 'wc-D-m2':'2026-06-13', 'wc-D-m3':'2026-06-19', 'wc-D-m4':'2026-06-19', 'wc-D-m5':'2026-06-26', 'wc-D-m6':'2026-06-26', 'wc-E-m1':'2026-06-14', 'wc-E-m2':'2026-06-14', 'wc-E-m3':'2026-06-20', 'wc-E-m4':'2026-06-21', 'wc-E-m5':'2026-06-25', 'wc-E-m6':'2026-06-25', 'wc-F-m1':'2026-06-14', 'wc-F-m2':'2026-06-15', 'wc-F-m3':'2026-06-20', 'wc-F-m4':'2026-06-20', 'wc-F-m5':'2026-06-25', 'wc-F-m6':'2026-06-25', 'wc-G-m1':'2026-06-15', 'wc-G-m2':'2026-06-16', 'wc-G-m3':'2026-06-21', 'wc-G-m4':'2026-06-22', 'wc-G-m5':'2026-06-27', 'wc-G-m6':'2026-06-27', 'wc-H-m1':'2026-06-15', 'wc-H-m2':'2026-06-16', 'wc-H-m3':'2026-06-21', 'wc-H-m4':'2026-06-22', 'wc-H-m5':'2026-06-27', 'wc-H-m6':'2026-06-27', 'wc-I-m1':'2026-06-16', 'wc-I-m2':'2026-06-17', 'wc-I-m3':'2026-06-22', 'wc-I-m4':'2026-06-23', 'wc-I-m5':'2026-06-26', 'wc-I-m6':'2026-06-26', 'wc-J-m1':'2026-06-16', 'wc-J-m2':'2026-06-17', 'wc-J-m3':'2026-06-22', 'wc-J-m4':'2026-06-23', 'wc-J-m5':'2026-06-28', 'wc-J-m6':'2026-06-28', 'wc-K-m1':'2026-06-17', 'wc-K-m2':'2026-06-18', 'wc-K-m3':'2026-06-23', 'wc-K-m4':'2026-06-24', 'wc-K-m5':'2026-06-27', 'wc-K-m6':'2026-06-27', 'wc-L-m1':'2026-06-17', 'wc-L-m2':'2026-06-17', 'wc-L-m3':'2026-06-23', 'wc-L-m4':'2026-06-23', 'wc-L-m5':'2026-06-27', 'wc-L-m6':'2026-06-27'}

# ── Persistent stores (SQLite-backed, survive restarts) ────────────────────
from storage import PersistentDict
import storage as _storage_mod
print(f'[STORAGE] Backend attivo: {_storage_mod.backend_name()}', flush=True)
USERS        = PersistentDict('users')        # email -> {pw}
PROFILES     = PersistentDict('profiles')      # email -> {nickname, avatar, created_at}
PREDICTIONS  = PersistentDict('predictions')   # email -> {matchId: {pick, score}}
KO_PRED      = PersistentDict('ko_pred')        # email -> {koMatchId: {score, adv}}
KO_SUBMITTED = PersistentDict('ko_submitted')   # email -> bool
SUBMITTED    = PersistentDict('submitted')      # email -> [group_letters]
RESULTS      = PersistentDict('results')        # matchId -> {pick, score}
MANUAL_RESULTS = PersistentDict('manual_results')  # matchId -> True (risultati bloccati dall'admin, il poll non li tocca)
RESET_TOKENS = PersistentDict('reset_tokens')   # token -> email
# Leagues
LEAGUES      = PersistentDict('leagues')        # league_id -> {name, password, admin_email, members:[], created_at}
USER_LEAGUES = PersistentDict('user_leagues')   # email -> [league_id, ...]
# Special predictions
TOPSCORER_PRED = PersistentDict('topscorer_pred')  # email -> player_name
FINAL_PRED     = PersistentDict('final_pred')      # email -> {home, away, winner, score}
# Real outcomes for special bets (admin sets these)
SPECIAL_RESULTS = PersistentDict('special_results')  # 'topscorer'->name, 'final'->{home,away,winner,score}
LIVE_CONFIG = PersistentDict('live_config')  # fonte live runtime: 'league_id','season','by_date' ('0'/'1')

def _live_cfg(key, default):
    v = LIVE_CONFIG.get(key)
    return v if v not in (None, '') else default
def _cfg_league_id():
    return str(_live_cfg('league_id', LIVE_LEAGUE_ID))
def _cfg_season():
    return str(_live_cfg('season', LIVE_SEASON))
def _cfg_by_date():
    raw = _live_cfg('by_date', '1' if os.environ.get('LIVE_BY_DATE', '1').strip() not in ('0', '', 'false', 'False') else '0')
    return str(raw).strip() not in ('0', '', 'false', 'False')

# ── Per-league prediction scoping ──────────────────────────────────────
# I pronostici sono salvati PER LEGA: la chiave dei vari store diventa
# "<league_id>|<email>", così ogni lega parte da zero e ha pronostici propri.
LK_SEP = '|'
def _lk(lid, email):
    return f"{lid}{LK_SEP}{email}"
def _split_lk(k):
    lid, _, em = k.partition(LK_SEP)
    return lid, em
def _req_league():
    """Estrae l'id lega dalla richiesta (query per GET, body per POST)."""
    if request.method == 'GET':
        return (request.args.get('league') or '').strip()
    return ((request.get_json(silent=True) or {}).get('league') or '').strip()
def _is_member(email, lid):
    lg = LEAGUES.get(lid)
    return bool(lg) and email in lg.get('members', [])


# Knockout match metadata for Excel export (id -> round label + matchup placeholder)
WC_KO_META = [
  ('wc-r64-73','Sedicesimi','2ª A - 2ª B'),('wc-r64-74','Sedicesimi','1ª E - 3ª'),
  ('wc-r64-75','Sedicesimi','1ª F - 2ª C'),('wc-r64-76','Sedicesimi','1ª C - 2ª F'),
  ('wc-r64-77','Sedicesimi','1ª I - 3ª'),('wc-r64-78','Sedicesimi','2ª E - 2ª I'),
  ('wc-r64-79','Sedicesimi','1ª A - 3ª'),('wc-r64-80','Sedicesimi','1ª L - 3ª'),
  ('wc-r64-81','Sedicesimi','1ª D - 3ª'),('wc-r64-82','Sedicesimi','1ª G - 3ª'),
  ('wc-r64-83','Sedicesimi','2ª K - 2ª L'),('wc-r64-84','Sedicesimi','1ª H - 2ª J'),
  ('wc-r64-85','Sedicesimi','1ª B - 3ª'),('wc-r64-86','Sedicesimi','1ª J - 2ª H'),
  ('wc-r64-87','Sedicesimi','1ª K - 3ª'),('wc-r64-88','Sedicesimi','2ª D - 2ª G'),
  ('wc-r32-89','Ottavi','V74 - V77'),('wc-r32-90','Ottavi','V73 - V75'),
  ('wc-r32-91','Ottavi','V76 - V78'),('wc-r32-92','Ottavi','V79 - V80'),
  ('wc-r32-93','Ottavi','V83 - V84'),('wc-r32-94','Ottavi','V81 - V82'),
  ('wc-r32-95','Ottavi','V86 - V88'),('wc-r32-96','Ottavi','V85 - V87'),
  ('wc-qf-97','Quarti','V89 - V90'),('wc-qf-98','Quarti','V93 - V94'),
  ('wc-qf-99','Quarti','V91 - V92'),('wc-qf-100','Quarti','V95 - V96'),
  ('wc-sf-101','Semifinale','V97 - V98'),('wc-sf-102','Semifinale','V99 - V100'),
  ('wc-bronze','Finale 3° posto','P SF1 - P SF2'),('wc-final','FINALE','V SF1 - V SF2'),
]

def hash_pw(pw):   return generate_password_hash(pw)
def league_pw_hash(pw): return hashlib.sha256(('lg:'+pw).encode()).hexdigest()
def verify_pw(stored, pw):
    """Verifica password: supporta nuovi hash werkzeug e vecchi SHA-256."""
    if not stored: return False
    if len(stored) == 64 and all(c in '0123456789abcdef' for c in stored.lower()):
        return hashlib.sha256(pw.encode()).hexdigest() == stored
    try:    return check_password_hash(stored, pw)
    except Exception: return False
def is_admin():    return session.get('email') == ADMIN_EMAIL
def deadline_passed(): return datetime.now(timezone.utc) > WC_DEADLINE

def _kickoff_map():
    """matchId -> datetime UTC del fischio d'inizio (per il blocco per-partita)."""
    m = {}
    for fr in FRIENDLY_SCHEDULE:
        if fr.get('kickoff'):
            try: m[fr['id']] = datetime.fromisoformat(fr['kickoff'].replace('Z','+00:00'))
            except Exception: pass
    return m

def match_locked(match_id):
    """Un pronostico è bloccato se:
       - la partita è iniziata (ora attuale >= kickoff), OPPURE
       - risulta già IN_PLAY/PAUSED/FINISHED dal feed live, OPPURE
       - per le partite del Mondiale, è passata la deadline globale."""
    now = datetime.now(timezone.utc)
    # Stato live (se l'abbiamo già rilevato)
    info = LIVE_STATE.get('matches', {}).get(match_id) if 'LIVE_STATE' in globals() else None
    if info and info.get('status') in ('IN_PLAY','PAUSED','FINISHED'):
        return True
    # Kickoff specifico (amichevoli)
    ko = _kickoff_map().get(match_id)
    if ko and now >= ko:
        return True
    # Partite del Mondiale: deadline globale
    if match_id.startswith('wc-') and deadline_passed():
        return True
    return False

def login_required(f):
    @wraps(f)
    def d(*a, **kw):
        if 'email' not in session: return jsonify({'error':'Non autenticato'}), 401
        return f(*a, **kw)
    return d

def admin_required(f):
    @wraps(f)
    def d(*a, **kw):
        if not is_admin(): return jsonify({'error':'Accesso negato'}), 403
        return f(*a, **kw)
    return d

# ── Rate limiting (in-memory, anti brute-force) ────────────────────────
_RATE = {}  # key (ip+route) -> [timestamps]
def rate_limit(max_calls, window_sec):
    def deco(f):
        @wraps(f)
        def d(*a, **kw):
            ip = request.headers.get('X-Forwarded-For', request.remote_addr or '?').split(',')[0].strip()
            key = f"{ip}:{f.__name__}"
            now = datetime.now(timezone.utc).timestamp()
            hits = [t for t in _RATE.get(key, []) if now - t < window_sec]
            if len(hits) >= max_calls:
                return jsonify({'error':'Troppi tentativi. Riprova tra qualche minuto.'}), 429
            hits.append(now)
            _RATE[key] = hits
            return f(*a, **kw)
        return d
    return deco

# ── Auth ───────────────────────────────────────────────────────────────
@app.route('/api/register', methods=['POST'])
@rate_limit(10, 600)
def register():
    d = request.json
    email    = d.get('email','').strip().lower()
    pw       = d.get('password','')
    nickname = d.get('nickname','').strip()
    avatar   = d.get('avatar','⚽')
    if not EMAIL_RE.match(email):
        return jsonify({'error':'Inserisci un indirizzo email valido'}), 400
    if len(pw) < 6:
        return jsonify({'error':'La password deve avere almeno 6 caratteri'}), 400
    if len(nickname) < 2 or len(nickname) > 24:
        return jsonify({'error':'Il nickname deve avere tra 2 e 24 caratteri'}), 400
    if email in USERS:
        return jsonify({'error':'Email già registrata'}), 409
    for e,p in PROFILES.items():
        if p.get('nickname','').lower() == nickname.lower():
            return jsonify({'error':'Nickname già in uso'}), 409
    now = datetime.now().strftime('%d/%m/%Y')
    USERS[email]   = {'pw': hash_pw(pw)}
    PROFILES[email]= {'nickname':nickname, 'avatar':avatar, 'created_at':now}
    USER_LEAGUES[email]= []
    session['email']   = email
    return jsonify({'ok':True,'email':email,'nickname':nickname,'avatar':avatar,
                    'is_admin':email==ADMIN_EMAIL,'created_at':now})

@app.route('/api/login', methods=['POST'])
@rate_limit(15, 300)
def login():
    d     = request.json
    email = d.get('email','').strip().lower()
    pw    = d.get('password','')
    if not email or not pw:
        return jsonify({'error':'Inserisci email e password.'}), 400
    if email not in USERS:
        return jsonify({'error':'Account non trovato. Registrati prima di accedere.'}), 404
    if not verify_pw(USERS[email].get('pw',''), pw):
        return jsonify({'error':'Password errata. Riprova.'}), 401
    # Migra al volo i vecchi hash SHA-256 al nuovo formato sicuro
    if len(USERS[email].get('pw','')) == 64:
        USERS[email]['pw'] = hash_pw(pw); USERS.save(email)
    session.permanent = True
    session['email'] = email
    p = PROFILES.get(email,{})
    return jsonify({'ok':True,'email':email,
                    'nickname':p.get('nickname',email.split('@')[0]),
                    'avatar':p.get('avatar','⚽'),
                    'created_at':p.get('created_at','—'),
                    'is_admin':email==ADMIN_EMAIL})

@app.route('/api/logout', methods=['POST'])
def logout():
    session.pop('email',None)
    return jsonify({'ok':True})

@app.route('/api/me')
def me():
    if 'email' not in session: return jsonify({'logged':False})
    email = session['email']
    p = PROFILES.get(email,{})
    return jsonify({'logged':True,'email':email,
                    'nickname':p.get('nickname',email.split('@')[0]),
                    'avatar':p.get('avatar','⚽'),
                    'created_at':p.get('created_at','—'),
                    'is_admin':email==ADMIN_EMAIL})

# ── Password reset ─────────────────────────────────────────────────────
@app.route('/api/request_reset', methods=['POST'])
@rate_limit(5, 600)
def request_reset():
    email = request.json.get('email','').strip().lower()
    if email not in USERS:
        return jsonify({'error':'Email non trovata'}), 404
    token = ''.join(secrets.choice(string.ascii_uppercase+string.digits) for _ in range(8))
    RESET_TOKENS[token] = email
    # Admin sees all tokens at /api/admin/reset_tokens
    return jsonify({'ok':True,'message':'Richiesta inviata. Contatta l\'amministratore con la tua email per ricevere il token.'})

@app.route('/api/reset_password', methods=['POST'])
def reset_password():
    d     = request.json
    token = d.get('token','').strip().upper()
    pw    = d.get('password','')
    if token not in RESET_TOKENS:
        return jsonify({'error':'Token non valido o scaduto'}), 400
    if len(pw) < 6:
        return jsonify({'error':'Password min 6 caratteri'}), 400
    email = RESET_TOKENS.pop(token)
    if email not in USERS:
        return jsonify({'error':'Account non trovato'}), 404
    USERS[email] = {'pw': hash_pw(pw)}
    return jsonify({'ok':True})

@app.route('/api/admin/reset_tokens')
@login_required
@admin_required
def admin_reset_tokens():
    return jsonify(dict(RESET_TOKENS))

# ── Predictions ────────────────────────────────────────────────────────
@app.route('/api/predictions', methods=['GET'])
@login_required
def get_predictions():
    email = session['email']
    lid = _req_league()
    if not lid or not _is_member(email, lid):
        return jsonify({'error':'Lega non valida'}), 400
    k = _lk(lid, email)
    return jsonify({'predictions':PREDICTIONS.get(k,{}),
                    'submitted':SUBMITTED.get(k,[]),
                    'topscorer':TOPSCORER_PRED.get(k,''),
                    'final_pred':FINAL_PRED.get(k,{}),
                    'ko_pred':KO_PRED.get(k,{}),
                    'ko_submitted':KO_SUBMITTED.get(k, False)})

@app.route('/api/ko_prediction', methods=['POST'])
@login_required
def set_ko_prediction():
    d        = request.json or {}
    match_id = (d.get('matchId') or '').strip()
    score    = (d.get('score') or '').strip()
    adv      = (d.get('adv') or '').strip()   # team that advances on a draw (supplementari/rigori)
    email    = session['email']
    lid      = _req_league()
    if not lid or not _is_member(email, lid):
        return jsonify({'error':'Lega non valida'}), 400
    k = _lk(lid, email)
    if not match_id:
        return jsonify({'error':'matchId mancante'}), 400
    if deadline_passed():
        return jsonify({'error':'Termine scaduto'}), 403
    if KO_SUBMITTED.get(k, False):
        return jsonify({'error':'Fase ad eliminazione già inviata e bloccata'}), 403
    if score:
        if '-' not in score:
            return jsonify({'error':'Formato risultato non valido'}), 400
        try:
            h, a = map(int, score.split('-', 1))
        except ValueError:
            return jsonify({'error':'Risultato non valido'}), 400
    if k not in KO_PRED:
        KO_PRED[k] = {}
    if score:
        entry = {'score': score}
        if adv:
            entry['adv'] = adv   # used only when the score is a draw
        KO_PRED[k][match_id] = entry
        KO_PRED.save(k)
    elif match_id in KO_PRED[k]:
        del KO_PRED[k][match_id]
        KO_PRED.save(k)
    return jsonify({'ok':True, 'ko_pred':KO_PRED[k]})

@app.route('/api/submit_ko', methods=['POST'])
@login_required
def submit_ko():
    email = session['email']
    lid   = _req_league()
    if not lid or not _is_member(email, lid):
        return jsonify({'error':'Lega non valida'}), 400
    if deadline_passed():
        return jsonify({'error':'Termine scaduto'}), 403
    KO_SUBMITTED[_lk(lid, email)] = True
    return jsonify({'ok':True, 'ko_submitted':True})


@app.route('/api/predictions', methods=['POST'])
@login_required
def set_prediction():
    d        = request.json
    match_id = d.get('matchId','').strip()
    pick     = d.get('pick','').strip()
    score    = d.get('score','').strip()
    email    = session['email']
    lid      = _req_league()
    if not lid or not _is_member(email, lid):
        return jsonify({'error':'Lega non valida'}), 400
    k = _lk(lid, email)

    if not match_id:
        return jsonify({'error':'matchId mancante'}), 400

    parts = match_id.split('-')
    group = parts[1].upper() if len(parts) >= 2 else None

    if group and group in SUBMITTED.get(k, []):
        return jsonify({'error': 'Pronostico già inviato per questo girone'}), 403

    # Blocco al fischio d'inizio: partita iniziata/in corso/finita = non più pronosticabile
    if match_locked(match_id):
        return jsonify({'error': 'La partita è già iniziata: pronostico bloccato'}), 403

    if k not in PREDICTIONS:
        PREDICTIONS[k] = {}

    # Derive pick from score if not provided
    if score and '-' in score and not pick:
        try:
            h, a = map(int, score.split('-', 1))
            pick = '1' if h > a else '2' if a > h else 'X'
        except ValueError:
            pass

    # Validate pick
    if pick not in ('1', 'X', '2'):
        return jsonify({'error': 'Pick non valido'}), 400

    # Save prediction — always overwrite with latest values
    PREDICTIONS[k][match_id] = {'pick': pick, 'score': score}
    PREDICTIONS.save(k)

    return jsonify({
        'ok': True,
        'predictions': PREDICTIONS[k],
        'submitted': SUBMITTED.get(k, [])
    })

@app.route('/api/submit_group', methods=['POST'])
@login_required
def submit_group():
    d     = request.json
    group = d.get('group','').upper()
    email = session['email']
    lid   = _req_league()
    if not lid or not _is_member(email, lid):
        return jsonify({'error':'Lega non valida'}), 400
    k = _lk(lid, email)
    if group not in WC_GROUPS:
        return jsonify({'error':'Girone non valido'}), 400
    if deadline_passed():
        return jsonify({'error':'Termine scaduto!'}), 403
    if group in SUBMITTED.get(k,[]):
        return jsonify({'error':'Già inviato'}), 409
    _sub = SUBMITTED.get(k, [])
    if group not in _sub: _sub.append(group)
    SUBMITTED[k] = _sub
    return jsonify({'ok':True,'submitted':SUBMITTED[k]})

# ── Top scorer prediction ──────────────────────────────────────────────
@app.route('/api/topscorer', methods=['POST'])
@login_required
def set_topscorer():
    player = request.json.get('player','').strip()
    email  = session['email']
    lid    = _req_league()
    if not lid or not _is_member(email, lid):
        return jsonify({'error':'Lega non valida'}), 400
    if not player: return jsonify({'error':'Giocatore non valido'}), 400
    if deadline_passed(): return jsonify({'error':'Termine scaduto'}), 403
    TOPSCORER_PRED[_lk(lid, email)] = player
    return jsonify({'ok':True,'player':player})

# ── Final prediction ───────────────────────────────────────────────────
@app.route('/api/final_pred', methods=['POST'])
@login_required
def set_final_pred():
    d = request.json
    home   = d.get('home','').strip()
    away   = d.get('away','').strip()
    winner = d.get('winner','').strip()
    score  = d.get('score','').strip()
    email  = session['email']
    lid    = _req_league()
    if not lid or not _is_member(email, lid):
        return jsonify({'error':'Lega non valida'}), 400
    if not home or not away or not winner:
        return jsonify({'error':'Dati non completi'}), 400
    if home == away:
        return jsonify({'error':'Le due finaliste devono essere diverse'}), 400
    if deadline_passed(): return jsonify({'error':'Termine scaduto'}), 403
    FINAL_PRED[_lk(lid, email)] = {'home':home,'away':away,'winner':winner,'score':score}
    return jsonify({'ok':True})

# ── Results (admin) ────────────────────────────────────────────────────
@app.route('/api/results', methods=['GET'])
def get_results():
    return jsonify(dict(RESULTS))

@app.route('/api/results', methods=['POST'])
@login_required
@admin_required
def set_result():
    """Imposta/corregge a mano il risultato di una partita. Diventa un OVERRIDE:
    il poll automatico non lo sovrascriverà più (finché non viene rilasciato)."""
    d = request.json
    mid   = d.get('matchId')
    score = (d.get('score') or '').strip()
    pick  = d.get('pick')
    # se arriva solo lo score (es. "3-0"), deduciamo l'esito
    if score and pick not in ('1','X','2'):
        try:
            h, a = map(int, score.split('-'))
            pick = '1' if h > a else '2' if a > h else 'X'
        except Exception:
            return jsonify({'error':'Punteggio non valido (usa formato "3-0")'}), 400
    if not mid or pick not in ('1','X','2'):
        return jsonify({'error':'Dati non validi'}), 400
    RESULTS[mid] = {'pick':pick,'score':score}
    MANUAL_RESULTS[mid] = True   # blocca: la fonte live non lo tocca più
    return jsonify({'ok':True, 'manual':True, 'result':{'pick':pick,'score':score}})

@app.route('/api/results/release', methods=['POST'])
@login_required
@admin_required
def release_result():
    """Rimuove l'override manuale: la partita torna a essere aggiornata dalla
    fonte live automatica al prossimo poll."""
    d = request.json or {}
    mid = d.get('matchId')
    if not mid:
        return jsonify({'error':'matchId mancante'}), 400
    if mid in MANUAL_RESULTS:
        MANUAL_RESULTS.pop(mid)
    LIVE_STATE['last_poll'] = 0   # forza un nuovo poll
    return jsonify({'ok':True, 'released':True})

@app.route('/api/manual_results', methods=['GET'])
@login_required
@admin_required
def get_manual_results():
    """Elenco dei matchId con override manuale attivo."""
    return jsonify({'manual': list(MANUAL_RESULTS.keys())})

@app.route('/api/special_results', methods=['GET'])
def get_special_results():
    return jsonify({'topscorer': SPECIAL_RESULTS.get('topscorer',''),
                    'final': SPECIAL_RESULTS.get('final',{})})

@app.route('/api/special_results', methods=['POST'])
@login_required
@admin_required
def set_special_results():
    """Admin imposta l'esito reale di capocannoniere e/o finale."""
    d = request.json or {}
    if 'topscorer' in d:
        SPECIAL_RESULTS['topscorer'] = (d.get('topscorer') or '').strip()
    if 'final' in d:
        f = d.get('final') or {}
        home = (f.get('home') or '').strip()
        away = (f.get('away') or '').strip()
        score = (f.get('score') or '').strip()
        winner = (f.get('winner') or '').strip()
        if home and away:
            if not winner and score and '-' in score:
                try:
                    h,a = map(int, score.split('-')); winner = home if h>=a else away
                except Exception: pass
            SPECIAL_RESULTS['final'] = {'home':home,'away':away,'winner':winner,'score':score}
    return jsonify({'ok':True,
                    'topscorer':SPECIAL_RESULTS.get('topscorer',''),
                    'final':SPECIAL_RESULTS.get('final',{})})

# ── Leagues ────────────────────────────────────────────────────────────
@app.route('/api/leagues', methods=['POST'])
@login_required
def create_league():
    d     = request.json
    name  = d.get('name','').strip()
    pw    = d.get('password','').strip()
    email = session['email']
    if not name or not pw:
        return jsonify({'error':'Nome e password obbligatori'}), 400
    # Il nome della lega deve essere univoco (confronto case-insensitive)
    for lg in LEAGUES.values():
        if lg.get('name','').strip().lower() == name.lower():
            return jsonify({'error':'Esiste già una lega con questo nome. Scegline un altro.'}), 409
    lid = secrets.token_urlsafe(8)
    LEAGUES[lid] = {
        'id':lid, 'name':name, 'password':league_pw_hash(pw),
        'admin_email':email, 'members':[email],
        'created_at':datetime.now().strftime('%d/%m/%Y')
    }
    _ul = USER_LEAGUES.get(email, []);  _ul.append(lid);  USER_LEAGUES[email] = _ul
    return jsonify({'ok':True,'league':_league_public(lid)})

@app.route('/api/leagues/join', methods=['POST'])
@login_required
def join_league():
    d    = request.json
    name = d.get('name','').strip().lower()
    pw   = d.get('password','').strip()
    email= session['email']
    target = None
    for lid, lg in LEAGUES.items():
        if lg['name'].lower() == name:
            target = lid; break
    if not target:
        return jsonify({'error':'Lega non trovata'}), 404
    if LEAGUES[target]['password'] != league_pw_hash(pw):
        return jsonify({'error':'Password errata'}), 401
    if email in LEAGUES[target]['members']:
        return jsonify({'error':'Sei già membro di questa lega'}), 409
    LEAGUES[target]['members'].append(email);  LEAGUES.save(target)
    _ul = USER_LEAGUES.get(email, []);  _ul.append(target);  USER_LEAGUES[email] = _ul
    return jsonify({'ok':True,'league':_league_public(target)})

@app.route('/api/leagues/join_by_link/<lid>', methods=['POST'])
@login_required
def join_by_link(lid):
    email = session['email']
    if lid not in LEAGUES:
        return jsonify({'error':'Lega non trovata'}), 404
    if email in LEAGUES[lid]['members']:
        return jsonify({'error':'Sei già membro'}), 409
    LEAGUES[lid]['members'].append(email);  LEAGUES.save(lid)
    _ul = USER_LEAGUES.get(email, []);  _ul.append(lid);  USER_LEAGUES[email] = _ul
    return jsonify({'ok':True,'league':_league_public(lid)})

@app.route('/api/leagues/mine')
@login_required
def my_leagues():
    email = session['email']
    ids = USER_LEAGUES.get(email,[])
    return jsonify([_league_public(lid) for lid in ids if lid in LEAGUES])

@app.route('/api/leagues/<lid>')
def league_detail(lid):
    if lid not in LEAGUES: return jsonify({'error':'Non trovata'}), 404
    lg = _league_public(lid)
    # Add leaderboard for this league
    lb = _league_leaderboard(lid)
    lg['leaderboard'] = lb
    return jsonify(lg)

def _league_public(lid):
    lg = LEAGUES[lid]
    return {
        'id': lg['id'], 'name': lg['name'],
        'admin_email': lg['admin_email'],
        'member_count': len(lg['members']),
        'created_at': lg['created_at'],
        'invite_link': f'/join/{lg["id"]}',
    }

def _league_leaderboard(lid):
    if lid not in LEAGUES: return []
    members = LEAGUES[lid]['members']
    board = []
    for email in members:
        p = PROFILES.get(email,{})
        pts, correct, exact = _calc_points(email, lid)
        board.append({'email':email,
                      'nickname':p.get('nickname',email.split('@')[0]),
                      'avatar':p.get('avatar','⚽'),
                      'points':pts,'correct':correct,'exact':exact,
                      'submitted':len(SUBMITTED.get(_lk(lid, email),[]))})
    board.sort(key=lambda x:-x['points'])
    return board

def _calc_points(email, lid):
    """Punteggio dell'utente IN UNA LEGA specifica.
    Categorie a punti (regolamento aggiornato):
      • Fase a gironi: risultato esatto = 3pt, solo esito 1/X/2 = 1pt
      • Capocannoniere indovinato = +5pt
      • Finale: entrambe le finaliste indovinate = 5pt, una sola = 3pt
    La fase a eliminazione diretta NON assegna punti.
    Ritorna (punti, esiti_corretti, risultati_esatti)."""
    pts=0; correct=0; exact=0
    k = _lk(lid, email)

    # 1) Gironi (unica fase di partite a punti)
    for mid, pred in PREDICTIONS.get(k, {}).items():
        res = RESULTS.get(mid)
        if not res: continue
        if pred.get('score') and pred.get('score') == res.get('score'):
            pts+=3; exact+=1; correct+=1
        elif pred.get('pick') == res.get('pick'):
            pts+=1; correct+=1

    # 2) Capocannoniere (+5)
    ts_real = SPECIAL_RESULTS.get('topscorer')
    if ts_real and TOPSCORER_PRED.get(k) == ts_real:
        pts += 5

    # 3) Finale: 2 finaliste giuste = 5pt, 1 sola = 3pt (nessun bonus vincitore)
    fin_real = SPECIAL_RESULTS.get('final')  # {home,away,winner,score}
    fp = FINAL_PRED.get(k)
    if fin_real and fp:
        real_teams = {t for t in (fin_real.get('home'), fin_real.get('away')) if t}
        pred_teams = {t for t in (fp.get('home'), fp.get('away')) if t}
        hit = len(real_teams & pred_teams)
        if hit >= 2:
            pts += 5
        elif hit == 1:
            pts += 3

    return pts, correct, exact


# ════════════════════════════════════════════════════════════════════════════
#  LIVE RESULTS ENGINE
#  - Con chiave: interroga football-data.org e aggiorna RESULTS dei match veri.
#  - Senza chiave: MODALITÀ SIMULAZIONE, fa progredire una partita demo nel tempo
#    così puoi vedere il meccanismo live senza registrarti.
# ════════════════════════════════════════════════════════════════════════════
import time as _time
try:
    import urllib.request as _urlreq
    import urllib.error as _urlerr
    import urllib.parse as _urlparse
except Exception:
    _urlreq = None
    _urlerr = None
    _urlparse = None

LIVE_STATE = {
    'last_poll': 0,            # timestamp ultimo polling
    'matches': {},            # matchId -> {home, away, score, status, minute}
    'error': None,
    'simulation': not LIVE_ENABLED,
}
# Memoria per la simulazione (avanzamento finto di una partita)
_SIM = {'start': None}

def _norm(name):
    """Normalizza un nome squadra per il confronto."""
    return (name or '').strip().lower()

def _alias(name):
    return LIVE_TEAM_ALIASES.get(name, name)

def _build_alias_lookup():
    """Mappa l'insieme {squadra1, squadra2} -> nostra partita, così riconosciamo
    la partita indipendentemente da quale squadra la fonte mette in casa.
    Conserva anche il nome normalizzato della NOSTRA squadra di casa, per
    orientare correttamente il punteggio."""
    lut = {}
    def add(mid, home, away):
        nh, na = _norm(_alias(home)), _norm(_alias(away))
        key = frozenset((nh, na))
        lut[key] = {'matchId': mid, 'home': home, 'away': away,
                    'home_norm': nh, 'away_norm': na}
    for grp in WC_MATCH_SCHEDULE.values():
        for m in grp:
            add(m['id'], m['home'], m['away'])
    for m in FRIENDLY_SCHEDULE:
        add(m['id'], m['home'], m['away'])
    return lut

# Calendario amichevoli lato server (specchio di data_friendlies.js) per il matching
FRIENDLY_SCHEDULE = [
    {'id':'fr-01','kickoff':'2026-06-09T23:00:00Z','home':'Senegal','away':'Arabia Saudita'},
    {'id':'fr-02','kickoff':'2026-06-10T01:00:00Z','home':'Argentina','away':'Islanda'},
    {'id':'fr-03','kickoff':'2026-06-10T01:00:00Z','home':'Iraq','away':'Venezuela'},
    {'id':'fr-04','kickoff':'2026-06-10T16:00:00Z','home':'Pakistan','away':'Afghanistan'},
    {'id':'fr-05','kickoff':'2026-06-10T19:45:00Z','home':'Portogallo','away':'Nigeria'},
]

def _hl_status(desc):
    """Mappa lo stato Highlightly (state.description) ai nostri 3 stati."""
    d = (desc or '').lower()
    if 'not started' in d or 'to be announced' in d or 'postponed' in d:
        return 'SCHEDULED'
    if any(k in d for k in ('finished', 'awarded')):
        return 'FINISHED'
    if any(k in d for k in ('first half','second half','half time','extra time',
                            'break time','penalties','in progress','suspended','interrupted')):
        return 'IN_PLAY'
    return 'SCHEDULED'

def _fetch_live_real():
    """Interroga Highlightly. Per le amichevoli (lega con centinaia di partite nel
    mondo) cerchiamo PER DATA, così becchiamo le nostre partite a prescindere
    dall'ordine. Per il Mondiale (poche partite) si usa leagueId via env.
    Doc: https://highlightly.net/football-api/documentation/
    """
    headers = _hl_headers()
    found = {}
    season = _cfg_season()

    # Modalità: se LIVE_BY_DATE attiva (default per le amichevoli), cerca per data.
    use_dates = _cfg_by_date()

    def _ingest(data):
        lut = _build_alias_lookup()
        for m in data.get('data', []):
            h = (m.get('homeTeam') or {}).get('name') or ''
            a = (m.get('awayTeam') or {}).get('name') or ''
            nh, na = _norm(h), _norm(a)
            match_ref = lut.get(frozenset((nh, na)))
            if not match_ref:
                continue
            st = m.get('state') or {}
            status = _hl_status(st.get('description'))
            cur = (st.get('score') or {}).get('current')
            score_str = ''
            if cur and '-' in cur:
                parts = [p.strip() for p in cur.split('-')]
                if len(parts) == 2 and parts[0].isdigit() and parts[1].isdigit():
                    api_home_goals, api_away_goals = parts[0], parts[1]
                    # Orienta il punteggio in base alla NOSTRA squadra di casa:
                    # se la casa secondo la fonte è la nostra casa, ordine invariato;
                    # altrimenti la fonte ha invertito e scambiamo.
                    if nh == match_ref['home_norm']:
                        score_str = f"{api_home_goals}-{api_away_goals}"
                    else:
                        score_str = f"{api_away_goals}-{api_home_goals}"
            found[match_ref['matchId']] = {
                'home': match_ref['home'], 'away': match_ref['away'],
                'score': score_str, 'status': status, 'minute': st.get('clock'),
            }

    def _get(url):
        req = _urlreq.Request(url, headers=headers)
        with _urlreq.urlopen(req, timeout=12) as resp:
            return json.loads(resp.read().decode('utf-8'))

    def _get_all_for_date(d):
        """Scarica le partite di una data scorrendo le pagine, ma si FERMA appena
        ha trovato tutte le nostre partite di quel giorno (per risparmiare richieste,
        il piano gratuito ne dà 100/giorno)."""
        # quante nostre partite ci aspettiamo in questa data (amichevoli + Mondiale)
        want = {fr['id'] for fr in FRIENDLY_SCHEDULE if (fr.get('kickoff') or '')[:10] == d}
        want |= {mid for mid, dd in WC_MATCH_DATES.items() if dd == d}
        offset = 0
        for _ in range(6):  # tetto di sicurezza: max 6 pagine
            data = _get(f"{HIGHLIGHTLY_BASE}/matches?date={d}&season={season}&limit=100&offset={offset}")
            _ingest(data)
            # se abbiamo già trovato tutte le nostre partite di oggi, basta
            if want and want.issubset(set(found.keys())):
                break
            pag = data.get('pagination') or {}
            total = pag.get('totalCount', 0)
            got = len(data.get('data', []))
            offset += got
            if got == 0 or offset >= total:
                break

    if use_dates:
        # Stessa logica delle amichevoli (ricerca PER DATA), estesa al Mondiale.
        # Per non sprecare le ~100 richieste/giorno del piano gratuito, interroghiamo
        # SOLO le date di OGGI/IERI (UTC) e solo se c'è almeno una NOSTRA partita di
        # quei giorni ancora SENZA risultato (in corso o appena finita).
        now_utc = datetime.now(timezone.utc)
        today = now_utc.date().isoformat()
        yest  = (now_utc.date() - timedelta(days=1)).isoformat()
        window = {today, yest}
        our_dates = {}
        for fr in FRIENDLY_SCHEDULE:
            if fr.get('kickoff'):
                our_dates[fr['id']] = fr['kickoff'][:10]
        for mid, dd in WC_MATCH_DATES.items():
            our_dates[mid] = dd
        pending_dates = set()
        for mid, dd in our_dates.items():
            if dd in window and mid not in RESULTS and mid not in MANUAL_RESULTS:
                pending_dates.add(dd)
        for d in sorted(pending_dates):
            try:
                _get_all_for_date(d)
            except Exception:
                continue
    else:
        # Mondiale: una lega con ~104 partite -> scorri le pagine (limit 100).
        league_id = _cfg_league_id()
        offset = 0
        for _ in range(6):  # tetto di sicurezza
            data = _get(f"{HIGHLIGHTLY_BASE}/matches?leagueId={league_id}&season={season}&limit=100&offset={offset}")
            _ingest(data)
            pag = data.get('pagination') or {}
            total = pag.get('totalCount', 0)
            got = len(data.get('data', []))
            offset += got
            if got == 0 or offset >= total:
                break

    return found

# Demo override: se SIM_DEMO è attivo, una partita "gioca" in 90 secondi reali
# per mostrare il meccanismo live anche fuori dagli orari veri delle partite.
SIM_DEMO = {'match_id': None, 'start': None}

def _sim_scripted(elapsed):
    if   elapsed < 20: return '0-0', 'IN_PLAY', max(1, int(elapsed*2))
    elif elapsed < 40: return '1-0', 'IN_PLAY', min(45, int(elapsed*1.2))
    elif elapsed < 60: return '1-1', 'IN_PLAY', min(70, int(elapsed*1.2))
    elif elapsed < 90: return '2-1', 'IN_PLAY', min(90, int(elapsed))
    else:              return '2-1', 'FINISHED', 90

def _fetch_live_sim():
    """Senza chiave reale NON inventiamo risultati: le amichevoli restano senza
    punteggio (l'admin può inserirlo a mano dal pannello). L'unica eccezione è
    SIM_DEMO, che l'admin attiva esplicitamente per vedere il meccanismo live."""
    now = datetime.now(timezone.utc)
    out = {}
    for fr in FRIENDLY_SCHEDULE:
        mid = fr['id']
        ko = fr.get('kickoff')
        ko_dt = None
        if ko:
            try: ko_dt = datetime.fromisoformat(ko.replace('Z','+00:00'))
            except Exception: ko_dt = None

        # Demo override: SOLO se l'admin l'ha avviata di proposito
        if SIM_DEMO['match_id'] == mid and SIM_DEMO['start'] is not None:
            el = _time.time() - SIM_DEMO['start']
            sc, stt, mn = _sim_scripted(el)
            out[mid] = {'home':fr['home'],'away':fr['away'],'score':sc,'status':stt,'minute':mn}
            continue

        if ko_dt is None:
            continue
        mins = (now - ko_dt).total_seconds() / 60.0
        if mins < 0:
            # non ancora iniziata
            out[mid] = {'home':fr['home'],'away':fr['away'],'score':'','status':'SCHEDULED','minute':None}
        elif mins < 110:
            # in corso: nessun punteggio finto, solo stato "in gioco"
            out[mid] = {'home':fr['home'],'away':fr['away'],'score':'','status':'IN_PLAY','minute':int(min(90,mins))}
        else:
            # finita: nessun risultato inventato. Resta da inserire a mano.
            out[mid] = {'home':fr['home'],'away':fr['away'],'score':'','status':'FINISHED','minute':90}
    return out

def poll_live(force=False):
    """Interroga la sorgente live (con cache di LIVE_POLL_SECONDS) e aggiorna RESULTS
    per le sole partite FINISHED (i punti contano solo a partita finita)."""
    now = _time.time()
    if not force and (now - LIVE_STATE['last_poll']) < LIVE_POLL_SECONDS:
        return LIVE_STATE
    LIVE_STATE['last_poll'] = now
    try:
        if LIVE_ENABLED and _urlreq is not None:
            matches = _fetch_live_real()
            LIVE_STATE['simulation'] = False
        else:
            matches = _fetch_live_sim()
            LIVE_STATE['simulation'] = True
        LIVE_STATE['matches'] = matches
        LIVE_STATE['error'] = None
        # Aggiorna RESULTS per le partite FINISHED (così la classifica si muove).
        # NON tocca le partite con risultato corretto a mano dall'admin (override).
        for mid, info in matches.items():
            if mid in MANUAL_RESULTS:
                continue  # risultato bloccato manualmente: la fonte non lo sovrascrive
            if info['status'] == 'FINISHED' and info['score']:
                h, a = map(int, info['score'].split('-'))
                pick = '1' if h > a else '2' if a > h else 'X'
                RESULTS[mid] = {'pick': pick, 'score': info['score']}
    except Exception as e:
        LIVE_STATE['error'] = str(e)
    return LIVE_STATE

@app.route('/api/live_raw')
@login_required
@admin_required
def api_live_raw():
    """[Diagnostica admin] cerca PER DATA (le date delle amichevoli) e mostra
    SOLO le partite che coinvolgono le nostre nazionali, con data/stato/punteggio
    grezzi da Highlightly. Serve a capire come l'API cataloga le partite."""
    if not LIVE_ENABLED:
        return jsonify({'error':'Chiave HIGHLIGHTLY_KEY non impostata'})
    headers = _hl_headers()
    # nomi (in inglese) delle nostre nazionali, per filtrare i risultati grezzi
    our_names = set()
    for fr in FRIENDLY_SCHEDULE:
        our_names.add(_norm(LIVE_TEAM_ALIASES.get(fr['home'], fr['home'])))
        our_names.add(_norm(LIVE_TEAM_ALIASES.get(fr['away'], fr['away'])))
    dates = sorted({fr['kickoff'][:10] for fr in FRIENDLY_SCHEDULE if fr.get('kickoff')})
    report = {}
    for d in dates:
        block = {'total_in_date': None, 'our_matches': []}
        try:
            offset = 0
            total = None
            for _ in range(6):
                url = f"{HIGHLIGHTLY_BASE}/matches?date={d}&season={LIVE_SEASON}&limit=100&offset={offset}"
                req = _urlreq.Request(url, headers=headers)
                with _urlreq.urlopen(req, timeout=12) as resp:
                    data = json.loads(resp.read().decode('utf-8'))
                total = (data.get('pagination') or {}).get('totalCount', 0)
                block['total_in_date'] = total
                got = len(data.get('data', []))
                for m in data.get('data', []):
                    h = (m.get('homeTeam') or {}).get('name') or ''
                    a = (m.get('awayTeam') or {}).get('name') or ''
                    if _norm(h) in our_names or _norm(a) in our_names:
                        st = m.get('state') or {}
                        block['our_matches'].append({
                            'home': h, 'away': a,
                            'status': st.get('description'),
                            'score': (st.get('score') or {}).get('current') or '',
                            'date': m.get('date'),
                            'league': (m.get('league') or {}).get('name'),
                        })
                offset += got
                if got == 0 or offset >= total:
                    break
        except Exception as e:
            block['error'] = str(e)
        report[d] = block
    return jsonify({'dates_queried': dates, 'report': report})

@app.route('/api/source_test')
@login_required
@admin_required
def api_source_test():
    """[Diagnostica admin] prova a contattare DA RENDER le varie fonti di dati
    e riporta cosa risponde ciascuna (200 = funziona, 403 = IP bloccato, ecc.).
    Serve a capire quale fonte è utilizzabile dal server."""
    def _probe(url, headers, label):
        out = {'label': label, 'url': url}
        try:
            req = _urlreq.Request(url, headers=headers)
            with _urlreq.urlopen(req, timeout=12) as resp:
                body = resp.read(400).decode('utf-8', 'replace')
                out['status'] = resp.status
                out['ok'] = True
                out['sample'] = body[:300]
        except _urlerr.HTTPError as e:
            out['status'] = e.code
            out['ok'] = False
            try: out['sample'] = e.read(300).decode('utf-8','replace')
            except Exception: out['sample'] = ''
        except Exception as e:
            out['status'] = None
            out['ok'] = False
            out['error'] = str(e)
        return out

    results = []

    # 1) Highlightly con entrambi gli header (come fa l'app ora)
    if HIGHLIGHTLY_KEY:
        results.append(_probe(
            f"{HIGHLIGHTLY_BASE}/matches?leagueId={LIVE_LEAGUE_ID}&season={LIVE_SEASON}&limit=3",
            {'x-rapidapi-key': HIGHLIGHTLY_KEY, 'x-rapidapi-host': HIGHLIGHTLY_HOST},
            'Highlightly (key+host)'))
        # 1b) Highlightly con un User-Agent "da browser" (a volte sblocca Cloudflare)
        results.append(_probe(
            f"{HIGHLIGHTLY_BASE}/matches?leagueId={LIVE_LEAGUE_ID}&season={LIVE_SEASON}&limit=3",
            {'x-rapidapi-key': HIGHLIGHTLY_KEY, 'x-rapidapi-host': HIGHLIGHTLY_HOST,
             'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36',
             'Accept': 'application/json'},
            'Highlightly (key+host+browser UA)'))
    else:
        results.append({'label':'Highlightly', 'error':'HIGHLIGHTLY_KEY non impostata'})

    # 2) football-data.org (se la chiave è presente)
    fd_key = os.environ.get('FOOTBALL_DATA_KEY', '')
    if fd_key:
        results.append(_probe(
            "https://api.football-data.org/v4/competitions/WC/matches",
            {'X-Auth-Token': fd_key},
            'football-data.org WC'))
    else:
        results.append({'label':'football-data.org', 'error':'FOOTBALL_DATA_KEY non impostata'})

    # 3) Test "nudo" verso un sito qualunque, per vedere se Render esce su internet
    results.append(_probe("https://api.github.com/zen", {'User-Agent':'toto-test'}, 'Internet check (github)'))

    return jsonify({'tested_from':'Render server', 'results': results})

@app.route('/api/live')
def api_live():
    """Stato live corrente: l'app frontend lo interroga periodicamente."""
    poll_live()
    kmap = _kickoff_map()
    now = datetime.now(timezone.utc)
    locked = {}
    kickoffs = {}
    for mid, ko in kmap.items():
        kickoffs[mid] = ko.isoformat()
        locked[mid] = match_locked(mid)
    return jsonify({
        'enabled': LIVE_ENABLED,
        'simulation': LIVE_STATE['simulation'],
        'matches': LIVE_STATE['matches'],
        'locked': locked,
        'kickoffs': kickoffs,
        'server_time': now.isoformat(),
        'error': LIVE_STATE['error'],
        'poll_seconds': LIVE_POLL_SECONDS,
    })

@app.route('/api/admin/live_config', methods=['GET'])
@login_required
@admin_required
def admin_live_config_get():
    return jsonify({
        'enabled': LIVE_ENABLED,
        'simulation': LIVE_STATE.get('simulation'),
        'league_id': _cfg_league_id(),
        'season': _cfg_season(),
        'by_date': _cfg_by_date(),
        'default_league_id': LIVE_LEAGUE_ID,
        'matches_found': len(LIVE_STATE.get('matches', {})),
        'error': LIVE_STATE.get('error'),
    })

@app.route('/api/admin/live_config', methods=['POST'])
@login_required
@admin_required
def admin_live_config_set():
    d = request.json or {}
    if 'league_id' in d:
        LIVE_CONFIG['league_id'] = str(d.get('league_id') or '').strip()
    if 'season' in d:
        LIVE_CONFIG['season'] = str(d.get('season') or '').strip()
    if 'by_date' in d:
        LIVE_CONFIG['by_date'] = '1' if d.get('by_date') else '0'
    # forza un poll subito per applicare e verificare
    LIVE_STATE['last_poll'] = 0
    poll_live(force=True)
    return jsonify({
        'ok': True,
        'league_id': _cfg_league_id(), 'season': _cfg_season(), 'by_date': _cfg_by_date(),
        'matches_found': len(LIVE_STATE.get('matches', {})),
        'matches': LIVE_STATE.get('matches', {}),
        'error': LIVE_STATE.get('error'),
    })

@app.route('/api/admin/live_test')
@login_required
@admin_required
def admin_live_test():
    """Forza un aggiornamento live adesso e riporta cosa è stato trovato."""
    LIVE_STATE['last_poll'] = 0
    poll_live(force=True)
    return jsonify({
        'enabled': LIVE_ENABLED,
        'simulation': LIVE_STATE.get('simulation'),
        'league_id': _cfg_league_id(), 'season': _cfg_season(), 'by_date': _cfg_by_date(),
        'matches': LIVE_STATE.get('matches', {}),
        'error': LIVE_STATE.get('error'),
    })

@app.route('/api/admin/import_predictions', methods=['POST'])
@login_required
@admin_required
def admin_import_predictions():
    """Ripristina/importa i pronostici dei gironi di un utente in una lega.
    Utile dopo il passaggio ai pronostici per-lega (vecchi pronostici globali)."""
    d = request.json or {}
    email = (d.get('email') or '').strip().lower()
    lid   = (d.get('league') or '').strip()
    preds = d.get('predictions') or {}
    mark_submitted = d.get('submitted', True)
    if email not in USERS:
        return jsonify({'error': f'Utente non registrato: {email}'}), 400
    if lid not in LEAGUES:
        return jsonify({'error': 'Lega non valida'}), 400
    if email not in LEAGUES[lid].get('members', []):
        return jsonify({'error': f'{email} non è membro della lega "{LEAGUES[lid]["name"]}"'}), 400
    if not isinstance(preds, dict) or not preds:
        return jsonify({'error': 'Nessun pronostico valido da importare'}), 400
    k = _lk(lid, email)
    cur = dict(PREDICTIONS.get(k, {}))
    groups = set(); count = 0
    for mid, val in preds.items():
        if not isinstance(val, dict):
            continue
        score = (val.get('score') or '').strip()
        pick  = (val.get('pick') or '').strip()
        if pick not in ('1', 'X', '2'):
            if '-' in score:
                try:
                    h, a = map(int, score.split('-', 1))
                    pick = '1' if h > a else '2' if a > h else 'X'
                except Exception:
                    continue
            else:
                continue
        cur[mid] = {'pick': pick, 'score': score}
        count += 1
        parts = mid.split('-')
        if len(parts) >= 2:
            groups.add(parts[1].upper())
    PREDICTIONS[k] = cur
    if mark_submitted and groups:
        sub = set(SUBMITTED.get(k, [])) | groups
        SUBMITTED[k] = sorted(sub)
    return jsonify({'ok': True, 'imported': count, 'groups': sorted(groups),
                    'email': email, 'league': LEAGUES[lid]['name']})

@app.route('/api/admin/league_members')
@login_required
@admin_required
def admin_league_members():
    lid = (request.args.get('league') or '').strip()
    if lid not in LEAGUES:
        return jsonify({'error': 'Lega non valida'}), 400
    lg = LEAGUES[lid]
    out = []
    for em in lg.get('members', []):
        p = PROFILES.get(em, {})
        out.append({'email': em,
                    'nickname': p.get('nickname', em.split('@')[0]),
                    'avatar': p.get('avatar', '⚽'),
                    'is_admin': em == lg.get('admin_email')})
    return jsonify({'league': lg['name'], 'admin_email': lg.get('admin_email'), 'members': out})

@app.route('/api/admin/remove_member', methods=['POST'])
@login_required
@admin_required
def admin_remove_member():
    d = request.json or {}
    email = (d.get('email') or '').strip().lower()
    lid   = (d.get('league') or '').strip()
    if lid not in LEAGUES:
        return jsonify({'error': 'Lega non valida'}), 400
    lg = LEAGUES[lid]
    if email not in lg.get('members', []):
        return jsonify({'error': 'Utente non presente in questa lega'}), 400
    if email == lg.get('admin_email'):
        return jsonify({'error': 'Non puoi rimuovere il creatore della lega'}), 400
    # rimuovi dalla lega
    lg['members'] = [m for m in lg['members'] if m != email]
    LEAGUES.save(lid)
    # rimuovi la lega dall'elenco dell'utente
    ul = USER_LEAGUES.get(email, [])
    if lid in ul:
        USER_LEAGUES[email] = [x for x in ul if x != lid]
    # cancella i suoi pronostici di QUESTA lega (gli altri restano)
    k = _lk(lid, email)
    for store in (PREDICTIONS, SUBMITTED, KO_PRED, KO_SUBMITTED, TOPSCORER_PRED, FINAL_PRED):
        if k in store:
            del store[k]
    return jsonify({'ok': True, 'email': email, 'league': lg['name'], 'members': len(lg['members'])})

@app.route('/api/admin/leagues_lookup')
@login_required
@admin_required
def admin_leagues_lookup():
    """[Diagnostica] cerca su Highlightly le leghe che corrispondono a una parola
    (es. 'World Cup'), per scoprire l'id della Coppa del Mondo da impostare."""
    if not LIVE_ENABLED:
        return jsonify({'error': 'Chiave HIGHLIGHTLY_KEY non impostata'}), 400
    if _urlreq is None:
        return jsonify({'error': 'Rete non disponibile'}), 400
    q = (request.args.get('q') or 'World Cup').strip()
    season = _cfg_season()
    headers = _hl_headers()
    out = []
    seen = set()
    tried = []
    # Prova alcune forme di query note dell'endpoint /leagues
    candidates = [
        f"{HIGHLIGHTLY_BASE}/leagues?leagueName={_urlparse.quote(q)}&limit=100",
        f"{HIGHLIGHTLY_BASE}/leagues?leagueName={_urlparse.quote(q)}&season={season}&limit=100",
        f"{HIGHLIGHTLY_BASE}/leagues?name={_urlparse.quote(q)}&limit=100",
    ]
    err = None
    for url in candidates:
        tried.append(url)
        try:
            req = _urlreq.Request(url, headers=headers)
            with _urlreq.urlopen(req, timeout=12) as resp:
                data = json.loads(resp.read().decode('utf-8'))
        except Exception as e:
            err = str(e)
            continue
        rows = data.get('data') if isinstance(data, dict) else data
        if not isinstance(rows, list):
            continue
        for lg in rows:
            lid = lg.get('id')
            if lid in seen:
                continue
            seen.add(lid)
            out.append({
                'id': lid,
                'name': lg.get('name'),
                'country': (lg.get('country') or {}).get('name') if isinstance(lg.get('country'), dict) else lg.get('country'),
                'logo': lg.get('logo'),
                'seasons': lg.get('seasons'),
            })
        if out:
            break
    return jsonify({'query': q, 'season': season, 'results': out, 'error': None if out else (err or 'Nessun risultato'), 'tried': tried})

@app.route('/api/sim_demo', methods=['POST'])
@login_required
@admin_required
def sim_demo():
    """[Solo test/simulazione] avvia la partita demo che 'gioca' in 90 secondi."""
    if LIVE_ENABLED:
        return jsonify({'error':'Demo non disponibile con dati reali attivi'}), 400
    d = request.json or {}
    mid = (d.get('matchId') or 'fr-01').strip()
    SIM_DEMO['match_id'] = mid
    SIM_DEMO['start'] = _time.time()
    LIVE_STATE['last_poll'] = 0
    poll_live(force=True)
    return jsonify({'ok':True, 'match_id':mid})


# ── Friendlies: scoring & leaderboard (test pre-Mondiale) ──────────────────
FRIENDLY_IDS = {m['id'] for m in FRIENDLY_SCHEDULE}

def _calc_friendly_points(email):
    preds = PREDICTIONS.get(email, {})
    pts=0; correct=0; exact=0
    for mid, pred in preds.items():
        if mid not in FRIENDLY_IDS:
            continue
        res = RESULTS.get(mid)
        if not res:
            continue
        score_ok = bool(pred.get('score')) and pred.get('score') == res.get('score')
        pick_ok  = pred.get('pick') == res.get('pick')
        if score_ok:   pts+=3; exact+=1; correct+=1
        elif pick_ok:  pts+=1; correct+=1
    return pts, correct, exact

@app.route('/api/friendly_leaderboard')
def friendly_leaderboard():
    board=[]
    for email in PREDICTIONS:
        has_fr = any(mid in FRIENDLY_IDS for mid in PREDICTIONS.get(email, {}))
        if not has_fr:
            continue
        p = PROFILES.get(email, {})
        pts, correct, exact = _calc_friendly_points(email)
        board.append({'email':email,
                      'nickname':p.get('nickname', email.split('@')[0]),
                      'avatar':p.get('avatar','⚽'),
                      'points':pts,'correct':correct,'exact':exact})
    board.sort(key=lambda x:(-x['points'], -x['exact']))
    return jsonify(board[:50])


# ── Global leaderboard ─────────────────────────────────────────────────
@app.route('/api/leaderboard')
def leaderboard():
    # Classifica globale deprecata: ora le classifiche sono solo per-lega.
    return jsonify([])

# ── Admin: elenco completo degli utenti registrati ────────────────────
@app.route('/api/admin/users')
@login_required
@admin_required
def admin_users():
    users = []
    for email in USERS:
        p = PROFILES.get(email, {})
        league_ids = [lid for lid in USER_LEAGUES.get(email, []) if lid in LEAGUES]
        pts = correct = exact = 0
        gp = kp = sg = 0; ko_sub = False
        for lid in league_ids:
            a,b,c = _calc_points(email, lid)
            pts+=a; correct+=b; exact+=c
            k=_lk(lid,email)
            gp += len(PREDICTIONS.get(k,{}))
            kp += len(KO_PRED.get(k,{}))
            sg += len(SUBMITTED.get(k,[]))
            ko_sub = ko_sub or bool(KO_SUBMITTED.get(k, False))
        leagues = [LEAGUES[lid]['name'] for lid in league_ids]
        users.append({
            'email': email,
            'nickname': p.get('nickname', email.split('@')[0]),
            'avatar': p.get('avatar', '⚽'),
            'created_at': p.get('created_at', '—'),
            'points': pts, 'correct': correct, 'exact': exact,
            'group_preds': gp,
            'ko_preds': kp,
            'submitted_groups': sg,
            'ko_submitted': ko_sub,
            'leagues': leagues,
            'is_admin': email == ADMIN_EMAIL,
        })
    users.sort(key=lambda u: u['nickname'].lower())
    return jsonify({'count': len(users), 'users': users})

# ── User profile ───────────────────────────────────────────────────────
@app.route('/api/league_predictions')
@login_required
def league_predictions():
    """Pronostici di un concorrente in una lega, visibili agli altri membri
    della stessa lega (solo dopo il termine, per evitare di copiare)."""
    me  = session['email']
    lid = (request.args.get('league') or '').strip()
    nick = (request.args.get('nick') or '').strip()
    if not lid or not _is_member(me, lid):
        return jsonify({'error': 'Lega non valida'}), 400
    target = None
    for em in LEAGUES[lid].get('members', []):
        if PROFILES.get(em, {}).get('nickname', '').lower() == nick.lower():
            target = em
            break
    if not target:
        return jsonify({'error': 'Concorrente non trovato in questa lega'}), 404
    p = PROFILES.get(target, {})
    pts, correct, exact = _calc_points(target, lid)
    k = _lk(lid, target)
    reveal = deadline_passed() or (target == me)
    return jsonify({
        'nickname': p.get('nickname', target.split('@')[0]),
        'avatar': p.get('avatar', '⚽'),
        'is_me': target == me,
        'locked': reveal,   # True = pronostici mostrati
        'points': pts, 'correct': correct, 'exact': exact,
        'submitted': len(SUBMITTED.get(k, [])),
        'predictions': PREDICTIONS.get(k, {}) if reveal else {},
        'topscorer': TOPSCORER_PRED.get(k, '') if reveal else '',
        'final_pred': FINAL_PRED.get(k, {}) if reveal else {},
    })

@app.route('/api/profile/<nickname>')
def get_profile(nickname):
    for email,p in PROFILES.items():
        if p.get('nickname','').lower() == nickname.lower():
            league_ids = [lid for lid in USER_LEAGUES.get(email,[]) if lid in LEAGUES]
            pts=correct=exact=0; sg=0
            for lid in league_ids:
                a,b,c=_calc_points(email, lid); pts+=a; correct+=b; exact+=c
                sg += len(SUBMITTED.get(_lk(lid,email),[]))
            leagues = [_league_public(lid) for lid in league_ids]
            return jsonify({'email':email,'nickname':p['nickname'],
                            'avatar':p.get('avatar','⚽'),
                            'created_at':p.get('created_at','—'),
                            'points':pts,'correct':correct,'exact':exact,
                            'submitted':sg,
                            'topscorer':'',
                            'final_pred':{},
                            'leagues':leagues})
    return jsonify({'error':'Profilo non trovato'}), 404

# ── Players API proxy ──────────────────────────────────────────────────
@app.route('/api/players/search')
@login_required
def search_players():
    query = request.args.get('q','').strip()
    if not query or len(query)<2:
        return jsonify({'players':[]})
    if not API_FOOTBALL_KEY:
        # Return static fallback if no key configured
        return jsonify({'players':[], 'error':'API key non configurata. Imposta API_FOOTBALL_KEY nel file config.py'})
    import requests as req
    try:
        resp = req.get('https://v3.football.api-sports.io/players',
            headers={'x-apisports-key': API_FOOTBALL_KEY},
            params={'search': query, 'league': 1, 'season': 2026},
            timeout=8)
        data = resp.json()
        players = []
        for item in data.get('response',[]):
            pl = item.get('player',{})
            st = item.get('statistics',[{}])[0]
            players.append({
                'id':       pl.get('id'),
                'name':     pl.get('name',''),
                'age':      pl.get('age',''),
                'nationality': pl.get('nationality',''),
                'photo':    pl.get('photo',''),
                'position': st.get('games',{}).get('position',''),
                'goals':    st.get('goals',{}).get('total',0) or 0,
                'assists':  st.get('goals',{}).get('assists',0) or 0,
                'games':    st.get('games',{}).get('appearences',0) or 0,
                'rating':   float(st.get('games',{}).get('rating',0) or 0),
            })
        return jsonify({'players': players})
    except Exception as e:
        return jsonify({'players':[], 'error': str(e)})

# ── Excel export ───────────────────────────────────────────────────────
@app.route('/api/export_excel')
@login_required
@admin_required
def export_excel():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    hdr_fill  = PatternFill("solid", fgColor="1A3A1A")
    hdr_font  = Font(bold=True, color="C8A44A", size=11)
    ttl_fill  = PatternFill("solid", fgColor="0A1A0A")
    ttl_font  = Font(bold=True, color="FFFFFF", size=13)
    # Row backgrounds: light, high-contrast so dark text is always readable
    white_fill = PatternFill("solid", fgColor="FFFFFF")
    alt_fill   = PatternFill("solid", fgColor="EDF3ED")   # very light green-grey
    ok_fill    = PatternFill("solid", fgColor="D6F5D6")   # light green (exact)
    ko_fill    = PatternFill("solid", fgColor="F9DADA")   # light red (wrong)
    # Dark text on light backgrounds
    data_font    = Font(color="1A1A1A", size=11)
    data_font_ok = Font(color="0A7A0A", size=11, bold=True)
    data_font_ko = Font(color="B01818", size=11)
    thin      = Side(style='thin', color="C8D4C8")
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)
    center    = Alignment(horizontal='center', vertical='center')
    wrap      = Alignment(horizontal='center', vertical='center', wrap_text=True)
    HEADERS   = ['Email','Nickname','Lega','Match ID','Partita','1/X/2','Punteggio','Ris. Reale','Score Reale','Punti','Inviato','Data Export']
    WIDTHS    = [28, 18, 18, 14, 34, 8, 14, 12, 12, 8, 10, 20]
    # Coppie (lega, email) da iterare: i pronostici sono per-lega
    member_pairs = []   # (lid, lname, email)
    for _lid, _lg in LEAGUES.items():
        for _em in _lg.get('members', []):
            member_pairs.append((_lid, _lg.get('name', _lid), _em))
    now_str   = datetime.now().strftime('%d/%m/%Y %H:%M')

    # Build match label map  id → "Casa – Ospite"
    match_labels = {}
    for grp_data in WC_MATCH_SCHEDULE.values():
        for m in grp_data:
            match_labels[m['id']] = f"{m['home']} – {m['away']}"

    for grp in WC_GROUPS:
        ws = wb.create_sheet(title=f"Girone {grp}")
        ws.merge_cells(f'A1:{get_column_letter(len(HEADERS))}1')
        c = ws['A1']
        c.value = f"TOTO CALCIO 2026 – Girone {grp}"
        c.fill = ttl_fill; c.font = ttl_font; c.alignment = center
        ws.row_dimensions[1].height = 28

        for col, (h, w) in enumerate(zip(HEADERS, WIDTHS), 1):
            c = ws.cell(row=2, column=col, value=h)
            c.fill = hdr_fill; c.font = hdr_font; c.alignment = center; c.border = border
            ws.column_dimensions[get_column_letter(col)].width = w

        row_i = 3
        # Include ogni (lega, utente) che ha pronostici per questo girone
        for lid, lname, email in sorted(member_pairs, key=lambda x:(x[1].lower(), x[2])):
            preds = PREDICTIONS.get(_lk(lid, email), {})
            grp_preds = {k: v for k, v in preds.items() if k.startswith(f'wc-{grp}-')}
            if not grp_preds:
                continue
            nick       = PROFILES.get(email, {}).get('nickname', email.split('@')[0])
            submitted  = grp in SUBMITTED.get(_lk(lid, email), [])

            for mid, pred in sorted(grp_preds.items()):
                res   = RESULTS.get(mid, {})
                pick  = pred.get('pick', '')
                score = pred.get('score', '')
                rp    = res.get('pick', '—')
                rs    = res.get('score', '—')
                label = match_labels.get(mid, mid)
                pts   = '—'
                row_fill = alt_fill if row_i % 2 == 0 else white_fill

                if res and pick:
                    if score and score == rs:
                        pts = 3
                        row_fill = ok_fill
                    elif pick == rp:
                        pts = 1
                    else:
                        pts = 0
                        row_fill = ko_fill

                values = [email, nick, lname, mid, label, pick or '—', score or '—',
                          rp, rs, pts, 'Sì' if submitted else 'No', now_str]
                # Pick font color based on outcome for readability
                if pts == 3:    cell_font = data_font_ok
                elif pts == 0:  cell_font = data_font_ko
                else:           cell_font = data_font
                for col, val in enumerate(values, 1):
                    c = ws.cell(row=row_i, column=col, value=val)
                    c.alignment = center; c.border = border; c.font = cell_font
                    c.fill = row_fill
                row_i += 1

        # Totals row per user per group
        ws.freeze_panes = 'A3'

    # ── Summary sheet ──────────────────────────────────────────────────────────
    ws_sum = wb.create_sheet(title="Classifica", index=0)
    ws_sum.merge_cells('A1:G1')
    c = ws_sum['A1']
    c.value = f"CLASSIFICHE PER LEGA – Esportato {now_str}"
    c.fill = ttl_fill; c.font = ttl_font; c.alignment = center
    ws_sum.row_dimensions[1].height = 28
    sum_headers = ['#', 'Lega', 'Email', 'Nickname', 'Punti', 'Corretti', 'Esatti']
    sum_widths  = [6, 18, 30, 20, 10, 12, 10]
    for col, (h, w) in enumerate(zip(sum_headers, sum_widths), 1):
        c = ws_sum.cell(row=2, column=col, value=h)
        c.fill = hdr_fill; c.font = hdr_font; c.alignment = center; c.border = border
        ws_sum.column_dimensions[get_column_letter(col)].width = w
    sum_rows = []
    for lid, lname, email in member_pairs:
        pts, correct, exact = _calc_points(email, lid)
        sum_rows.append((lname, pts, correct, exact, email))
    sum_rows.sort(key=lambda x: (x[0].lower(), -x[1], -x[2]))
    for i, (lname, pts, correct, exact, email) in enumerate(sum_rows, 1):
        nick = PROFILES.get(email, {}).get('nickname', email.split('@')[0])
        row  = [i, lname, email, nick, pts, correct, exact]
        fill = alt_fill if i % 2 == 0 else white_fill
        for col, val in enumerate(row, 1):
            c = ws_sum.cell(row=i+2, column=col, value=val)
            c.alignment = center; c.border = border; c.font = data_font
            c.fill = fill
    ws_sum.freeze_panes = 'A3'

    # ── Special predictions sheet ──────────────────────────────────────────────
    ws_sp = wb.create_sheet(title="Pronostici Speciali")
    ws_sp.merge_cells('A1:G1')
    c = ws_sp['A1']
    c.value = "PRONOSTICI SPECIALI – Capocannoniere & Finale"
    c.fill = ttl_fill; c.font = ttl_font; c.alignment = center
    sp_headers = ['Lega', 'Email', 'Nickname', 'Capocannoniere', 'Finaliste', 'Risultato Finale', 'Vincitore']
    sp_widths  = [18, 28, 18, 22, 36, 20, 18]
    for col, (h, w) in enumerate(zip(sp_headers, sp_widths), 1):
        c = ws_sp.cell(row=2, column=col, value=h)
        c.fill = hdr_fill; c.font = hdr_font; c.alignment = center; c.border = border
        ws_sp.column_dimensions[get_column_letter(col)].width = w
    row_i = 3
    for lid, lname, email in sorted(member_pairs, key=lambda x:(x[1].lower(), x[2])):
        k = _lk(lid, email)
        ts = TOPSCORER_PRED.get(k, '')
        fp = FINAL_PRED.get(k, {})
        if not ts and not fp:
            continue
        nick = PROFILES.get(email, {}).get('nickname', email.split('@')[0])
        finaliste = f"{fp.get('home','?')} vs {fp.get('away','?')}" if fp else '—'
        fscore    = fp.get('score', '—') if fp else '—'
        fwin      = fp.get('winner', '—') if fp else '—'
        values = [lname, email, nick, ts or '—', finaliste, fscore, fwin]
        fill = alt_fill if row_i % 2 == 0 else white_fill
        for col, val in enumerate(values, 1):
            c = ws_sp.cell(row=row_i, column=col, value=val)
            c.alignment = center; c.border = border; c.font = data_font
            c.fill = fill
        row_i += 1

    # ── Knockout phase sheet (user's bracket predictions) ──────────────────────
    ws_ko = wb.create_sheet(title="Fase Eliminazione")
    ws_ko.merge_cells('A1:G1')
    c = ws_ko['A1']
    c.value = "FASE AD ELIMINAZIONE DIRETTA – Pronostici (senza punti)"
    c.fill = ttl_fill; c.font = ttl_font; c.alignment = center
    ws_ko.row_dimensions[1].height = 28
    ko_headers = ['Lega', 'Email', 'Nickname', 'Turno', 'Sfida', 'Risultato', 'Qualificata (suppl./rig.)']
    ko_widths  = [18, 26, 16, 16, 20, 12, 24]
    for col, (h, w) in enumerate(zip(ko_headers, ko_widths), 1):
        c = ws_ko.cell(row=2, column=col, value=h)
        c.fill = hdr_fill; c.font = hdr_font; c.alignment = center; c.border = border
        ws_ko.column_dimensions[get_column_letter(col)].width = w
    row_i = 3
    for lid, lname, email in sorted(member_pairs, key=lambda x:(x[1].lower(), x[2])):
        kp = KO_PRED.get(_lk(lid, email), {})
        if not kp:
            continue
        nick = PROFILES.get(email, {}).get('nickname', email.split('@')[0])
        for mid, label, matchup in WC_KO_META:
            pred = kp.get(mid)
            if not pred:
                continue
            score = pred.get('score', '—')
            adv   = pred.get('adv', '')
            adv_disp = adv if adv else '—'
            values = [lname, email, nick, label, matchup, score, adv_disp]
            fill = alt_fill if row_i % 2 == 0 else white_fill
            for col, val in enumerate(values, 1):
                c = ws_ko.cell(row=row_i, column=col, value=val)
                c.alignment = center; c.border = border; c.font = data_font
                c.fill = fill
            row_i += 1
    ws_ko.freeze_panes = 'A3'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, download_name='toto_calcio_mondiale_2026.xlsx', as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ── Deadline ───────────────────────────────────────────────────────────
@app.route('/api/deadline')
def deadline_info():
    now=datetime.now(timezone.utc); passed=now>WC_DEADLINE
    return jsonify({'passed':passed,'seconds_remaining':max(0,int((WC_DEADLINE-now).total_seconds()))})

# ── Join via link (page) ───────────────────────────────────────────────
@app.route('/join/<lid>')
def join_page(lid):
    return render_template('index.html')

@app.route('/')
def index():
    return render_template('index.html')

# ── Sicurezza: header HTTP su ogni risposta ────────────────────────────
@app.after_request
def _security_headers(resp):
    # Impedisce di indovinare il MIME type
    resp.headers['X-Content-Type-Options'] = 'nosniff'
    # Niente embedding in iframe altrui (anti clickjacking)
    resp.headers['X-Frame-Options'] = 'SAMEORIGIN'
    # Non perdere il referer verso siti esterni
    resp.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
    # Limita le feature del browser
    resp.headers['Permissions-Policy'] = 'geolocation=(), microphone=(), camera=()'
    # In produzione (HTTPS) attiva HSTS: forza https per 1 anno
    if _is_prod:
        resp.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains'
    return resp

# Forza HTTPS in produzione (se il proxy segnala connessione non sicura)
@app.before_request
def _force_https():
    if _is_prod and not request.is_secure:
        # X-Forwarded-Proto viene gestito da ProxyFix; se ancora http → redirect
        proto = request.headers.get('X-Forwarded-Proto', 'http')
        if proto != 'https':
            from flask import redirect
            return redirect(request.url.replace('http://', 'https://', 1), code=301)

# ── Health check (per uptime monitor / load balancer) ──────────────────
@app.route('/healthz')
def _healthz():
    return jsonify({'status':'ok'}), 200

# ── Error handlers ─────────────────────────────────────────────────────
@app.errorhandler(404)
def _not_found(e):
    if request.path.startswith('/api/'):
        return jsonify({'error':'Risorsa non trovata'}), 404
    return render_template('index.html'), 200

@app.errorhandler(413)
def _too_large(e):
    return jsonify({'error':'Richiesta troppo grande (max 1MB)'}), 413

@app.errorhandler(500)
def _server_error(e):
    return jsonify({'error':'Errore interno del server'}), 500

if __name__=='__main__':
    # debug SOLO in sviluppo; in produzione si usa gunicorn (vedi Procfile)
    debug = os.environ.get('FLASK_DEBUG') == '1'
    port  = int(os.environ.get('PORT', 5001))
    app.run(host='0.0.0.0', port=port, debug=debug)